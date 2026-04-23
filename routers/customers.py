"""Customer router: CRUD + Excel import."""

import io
from datetime import date, datetime
from typing import Any, Dict, Optional
from sqlalchemy.exc import IntegrityError

from fastapi import APIRouter, Depends, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import JSONResponse, RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy import or_
from sqlalchemy.orm import Session

from database import get_db
from models import Customer

router = APIRouter()
templates = Jinja2Templates(directory="frontend/templates")

EXCEL_COLUMNS = ["ho_ten", "gioi_tinh", "ngay_sinh", "ngay_chet", "so_giay_to", "ngay_cap", "dia_chi"]
DATE_FIELDS = {"ngay_sinh", "ngay_chet", "ngay_cap"}


def parse_date(value: Any, allow_year_only: bool = True) -> Optional[date]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    if isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            from openpyxl.utils.datetime import from_excel

            dt = from_excel(value)
            if isinstance(dt, datetime):
                return dt.date()
            if isinstance(dt, date):
                return dt
        except Exception:
            pass

    s = str(value).strip()
    if not s or s.lower() in ("nan", "none"):
        return None

    if allow_year_only and len(s) == 4 and s.isdigit():
        y = int(s)
        if 1 <= y <= 9999:
            return date(y, 1, 1)

    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def format_date_display(d: Optional[date]) -> str:
    if not d:
        return ""
    if d.day == 1 and d.month == 1:
        return f"{d.year:04d}"
    return d.strftime("%d/%m/%Y")


def normalize_gender(value: str) -> str:
    s = (value or "").strip().lower()
    if not s:
        return ""
    if s in ("nam", "male", "m"):
        return "Nam"
    if s in ("nữ", "nu", "female", "f"):
        return "Nữ"
    return ""


def as_input_value(value: Any, is_date: bool = False) -> str:
    if value is None:
        return ""
    if is_date:
        d = parse_date(value, allow_year_only=True)
        if d:
            return format_date_display(d)
    return str(value).strip()


def to_customer_json(c: Customer) -> Dict[str, Any]:
    return {
        "id": c.id,
        "ho_ten": c.ho_ten,
        "gioi_tinh": c.gioi_tinh,
        "ngay_sinh": format_date_display(c.ngay_sinh),
        "ngay_chet": format_date_display(c.ngay_chet),
        "so_giay_to": c.so_giay_to,
        "ngay_cap": format_date_display(c.ngay_cap),
        "dia_chi": c.dia_chi,
    }


def validate_customer_form(form: Dict[str, str], db: Session, current_id: Optional[int] = None, duplicate_as_error: bool = True):
    field_errors: Dict[str, str] = {}

    required = ["ho_ten"]
    for key in required:
        if not (form.get(key) or "").strip():
            field_errors[key] = "Bat buoc"

    gioi_tinh = normalize_gender(form.get("gioi_tinh", ""))
    if form.get("gioi_tinh") and not gioi_tinh:
        field_errors["gioi_tinh"] = "Chon Nam hoac Nu"

    ngay_sinh = parse_date(form.get("ngay_sinh"), allow_year_only=True)
    ngay_chet = parse_date(form.get("ngay_chet"), allow_year_only=True)
    ngay_cap = parse_date(form.get("ngay_cap"), allow_year_only=True)

    if form.get("ngay_sinh") and ngay_sinh is None:
        field_errors["ngay_sinh"] = "Ngay khong hop le (yyyy hoac dd/mm/yyyy)"
    if form.get("ngay_chet") and ngay_chet is None:
        field_errors["ngay_chet"] = "Ngay khong hop le (yyyy hoac dd/mm/yyyy)"
    if form.get("ngay_cap") and ngay_cap is None:
        field_errors["ngay_cap"] = "Ngay khong hop le (yyyy hoac dd/mm/yyyy)"

    so_gt = (form.get("so_giay_to") or "").strip()
    if so_gt and duplicate_as_error:
        q = db.query(Customer).filter(Customer.so_giay_to == so_gt)
        if current_id is not None:
            q = q.filter(Customer.id != current_id)
        if q.first():
            field_errors["so_giay_to"] = "So giay to da ton tai"

    cleaned = {
        "ho_ten": (form.get("ho_ten") or "").strip(),
        "gioi_tinh": gioi_tinh,
        "ngay_sinh": ngay_sinh,
        "ngay_chet": ngay_chet,
        "so_giay_to": so_gt,
        "ngay_cap": ngay_cap,
        "dia_chi": (form.get("dia_chi") or "").strip(),
    }
    return cleaned, field_errors


def result_message(field_errors: Dict[str, str]) -> str:
    order = ["ho_ten", "gioi_tinh", "ngay_sinh", "ngay_chet", "so_giay_to", "ngay_cap", "dia_chi"]
    labels = {
        "ho_ten": "Thieu ho ten",
        "gioi_tinh": "Gioi tinh khong hop le",
        "ngay_sinh": "Ngay sinh khong hop le",
        "ngay_chet": "Ngay chet khong hop le",
        "so_giay_to": "So giay to khong hop le",
        "ngay_cap": "Ngay cap khong hop le",
        "dia_chi": "Thieu dia chi",
    }
    msg = [labels[k] for k in order if k in field_errors]
    return " | ".join(msg) if msg else "Du lieu khong hop le"


@router.get("/")
def list_customers(request: Request, db: Session = Depends(get_db), q: str = ""):
    query = db.query(Customer)
    if q:
        query = query.filter(or_(Customer.ho_ten.contains(q), Customer.so_giay_to.contains(q), Customer.dia_chi.contains(q)))
    customers = query.order_by(Customer.ho_ten).all()
    return templates.TemplateResponse("customers/list.html", {"request": request, "customers": customers, "q": q})


@router.get("/api/search")
def search_customers(db: Session = Depends(get_db), q: str = "", limit: int = 10):
    query = db.query(Customer)
    if q:
        query = query.filter(or_(
            Customer.ho_ten.contains(q), 
            Customer.so_giay_to.contains(q), 
            Customer.dia_chi.contains(q)
        ))
    customers = query.order_by(Customer.ho_ten).limit(limit).all()
    return JSONResponse({
        "ok": True, 
        "data": [to_customer_json(c) for c in customers]
    })


@router.get("/download-template")
def download_template():
    import openpyxl
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Danh sach nguoi"

    widths = [24, 14, 16, 16, 22, 16, 40]
    for idx, field in enumerate(EXCEL_COLUMNS, start=1):
        cell = ws.cell(row=1, column=idx, value=field)
        cell.font = Font(bold=True)
        ws.column_dimensions[get_column_letter(idx)].width = widths[idx - 1]

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=mau_nhap_nguoi.xlsx"},
    )


@router.post("/upload-excel")
async def upload_excel(request: Request, file: UploadFile = File(...), db: Session = Depends(get_db)):
    import openpyxl

    if not file.filename.endswith((".xlsx", ".xls")):
        return templates.TemplateResponse("customers/upload_result.html", {
            "request": request, "error_global": "Chi chap nhan file .xlsx",
            "results": [], "added": 0, "skipped": 0, "errors": 0, "total": 0
        })

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        return templates.TemplateResponse("customers/upload_result.html", {
            "request": request, "error_global": f"Khong the mo file: {e}",
            "results": [], "added": 0, "skipped": 0, "errors": 0, "total": 0
        })

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return templates.TemplateResponse("customers/upload_result.html", {
            "request": request, "error_global": "File khong co du lieu.",
            "results": [], "added": 0, "skipped": 0, "errors": 0, "total": 0
        })

    raw_headers = [str(h).strip() if h else "" for h in rows[0]]

    def find_col(keywords):
        for kw in keywords:
            for i, h in enumerate(raw_headers):
                if kw.lower() in h.lower():
                    return i
        return None

    col = {
        "ho_ten": find_col(["ho_ten", "h?", "ten", "full_name"]),
        "gioi_tinh": find_col(["gioi_tinh", "gioi", "gender"]),
        "ngay_sinh": find_col(["ngay_sinh", "ngay sinh", "birth"]),
        "ngay_chet": find_col(["ngay_chet", "ngay mat", "death", "chet"]),
        "so_giay_to": find_col(["so_giay_to", "cccd", "giay to", "id_number", "khai tu"]),
        "ngay_cap": find_col(["ngay_cap", "ngay cap", "issue"]),
        "dia_chi": find_col(["dia_chi", "dia chi", "address"]),
    }

    missing = [f for f in ["ho_ten"] if col[f] is None]
    if missing:
        return templates.TemplateResponse("customers/upload_result.html", {
            "request": request,
            "error_global": f"Khong nhan dien duoc cot: {', '.join(missing)}",
            "results": [], "added": 0, "skipped": 0, "errors": 0, "total": 0
        })

    def get_raw(row_vals, field):
        idx = col.get(field)
        if idx is None or idx >= len(row_vals):
            return None
        return row_vals[idx]

    results = []
    added_customers = []
    added = skipped = errors = 0

    for row_num, row in enumerate(rows[1:], start=2):
        raw_form = {
            "ho_ten": as_input_value(get_raw(row, "ho_ten")),
            "gioi_tinh": as_input_value(get_raw(row, "gioi_tinh")),
            "ngay_sinh": as_input_value(get_raw(row, "ngay_sinh"), is_date=True),
            "ngay_chet": as_input_value(get_raw(row, "ngay_chet"), is_date=True),
            "so_giay_to": as_input_value(get_raw(row, "so_giay_to")),
            "ngay_cap": as_input_value(get_raw(row, "ngay_cap"), is_date=True),
            "dia_chi": as_input_value(get_raw(row, "dia_chi")),
        }

        if not raw_form["ho_ten"] and not raw_form["so_giay_to"]:
            continue

        if raw_form["so_giay_to"] and db.query(Customer).filter(Customer.so_giay_to == raw_form["so_giay_to"]).first():
            results.append({"row": row_num, "name": raw_form["ho_ten"] or "?", "status": "skip", "message": "Trung so giay to, bo qua"})
            skipped += 1
            continue

        cleaned, field_errors = validate_customer_form(raw_form, db, duplicate_as_error=False)
        if field_errors:
            results.append({
                "row": row_num,
                "name": raw_form["ho_ten"] or "?",
                "status": "error",
                "message": result_message(field_errors),
                "raw": raw_form,
                "field_errors": field_errors,
            })
            errors += 1
            continue

        try:
            c = Customer(**cleaned)
            db.add(c)
            db.commit()
            db.refresh(c)
            added_customers.append(to_customer_json(c))
            results.append({"row": row_num, "name": cleaned["ho_ten"], "status": "ok", "message": "Them thanh cong"})
            added += 1
        except Exception as e:
            db.rollback()
            results.append({
                "row": row_num,
                "name": raw_form["ho_ten"] or "?",
                "status": "error",
                "message": f"Loi: {e}",
                "raw": raw_form,
                "field_errors": {},
            })
            errors += 1

    return templates.TemplateResponse("customers/upload_result.html", {
        "request": request,
        "error_global": None,
        "results": results,
        "added_customers": added_customers,
        "added": added,
        "skipped": skipped,
        "errors": errors,
        "total": added + skipped + errors,
    })


@router.post("/upload-excel/save-row")
def upload_excel_save_row(
    ho_ten: Optional[str] = Form(None),
    gioi_tinh: Optional[str] = Form(None),
    ngay_sinh: Optional[str] = Form(None),
    ngay_chet: Optional[str] = Form(None),
    so_giay_to: Optional[str] = Form(None),
    ngay_cap: Optional[str] = Form(None),
    dia_chi: Optional[str] = Form(None),
    db: Session = Depends(get_db),
):
    form = {
        "ho_ten": (ho_ten or "").strip(),
        "gioi_tinh": (gioi_tinh or "").strip(),
        "ngay_sinh": (ngay_sinh or "").strip(),
        "ngay_chet": (ngay_chet or "").strip(),
        "so_giay_to": (so_giay_to or "").strip(),
        "ngay_cap": (ngay_cap or "").strip(),
        "dia_chi": (dia_chi or "").strip(),
    }
    cleaned, field_errors = validate_customer_form(form, db, duplicate_as_error=True)
    if field_errors:
        return JSONResponse({"ok": False, "errors": field_errors, "message": result_message(field_errors)}, status_code=400)

    c = Customer(**cleaned)
    db.add(c)
    db.commit()
    db.refresh(c)
    return JSONResponse({"ok": True, "message": "Da luu thanh cong", "customer": to_customer_json(c)})


@router.get("/create")
def create_form(request: Request):
    form = {k: "" for k in EXCEL_COLUMNS}
    return templates.TemplateResponse("customers/form.html", {
        "request": request, "obj": None, "errors": [], "field_errors": {}, "form": form
    })


@router.post("/inline-create")
def inline_create(
    ho_ten: Optional[str] = Form(None), gioi_tinh: Optional[str] = Form(None),
    ngay_sinh: Optional[str] = Form(None), ngay_chet: Optional[str] = Form(None),
    so_giay_to: Optional[str] = Form(None), ngay_cap: Optional[str] = Form(None),
    dia_chi: Optional[str] = Form(None),
    db: Session = Depends(get_db)
):
    name = (ho_ten or "").strip()
    if not name:
        return JSONResponse({"ok": False, "errors": {"ho_ten": "Bat buoc"}}, status_code=400)

    so_giay_to_val = (so_giay_to or "").strip() or None
    # Upsert: nếu CCCD đã tồn tại → cập nhật với dữ liệu mới rồi trả về
    if so_giay_to_val:
        existing = db.query(Customer).filter(Customer.so_giay_to == so_giay_to_val).first()
        if existing:
            if name: existing.ho_ten = name
            gt = normalize_gender(gioi_tinh or "")
            if gt: existing.gioi_tinh = gt
            ns = parse_date(ngay_sinh or "", allow_year_only=True)
            if ns: existing.ngay_sinh = ns
            nc = parse_date(ngay_cap or "", allow_year_only=True)
            if nc: existing.ngay_cap = nc
            dc = (dia_chi or "").strip()
            if dc: existing.dia_chi = dc
            db.commit(); db.refresh(existing)
            return JSONResponse({"ok": True, "customer": to_customer_json(existing), "updated": True})

    c = Customer(
        ho_ten=name,
        gioi_tinh=normalize_gender(gioi_tinh or "") or None,
        ngay_sinh=parse_date(ngay_sinh or "", allow_year_only=True),
        ngay_chet=parse_date(ngay_chet or "", allow_year_only=True),
        so_giay_to=so_giay_to_val,
        ngay_cap=parse_date(ngay_cap or "", allow_year_only=True),
        dia_chi=(dia_chi or "").strip() or None,
    )
    db.add(c)
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        existing = db.query(Customer).filter(Customer.so_giay_to == so_giay_to_val).first()
        if existing:
            return JSONResponse({"ok": True, "customer": to_customer_json(existing)})
        return JSONResponse({"ok": False, "errors": {"so_giay_to": "Trùng"}}, status_code=400)
    db.refresh(c)
    return JSONResponse({"ok": True, "customer": to_customer_json(c)})


@router.post("/{cid}/quick-update")
def quick_update(
    cid: int,
    ho_ten: Optional[str] = Form(None), gioi_tinh: Optional[str] = Form(None),
    ngay_sinh: Optional[str] = Form(None), ngay_chet: Optional[str] = Form(None), so_giay_to: Optional[str] = Form(None),
    ngay_cap: Optional[str] = Form(None), dia_chi: Optional[str] = Form(None),
    db: Session = Depends(get_db)
):
    c = db.get(Customer, cid)
    if not c:
        raise HTTPException(status_code=404, detail="Not found")
    if ho_ten and ho_ten.strip():    c.ho_ten    = ho_ten.strip()
    gt = normalize_gender(gioi_tinh or "")
    if gt:                           c.gioi_tinh = gt
    ns = parse_date(ngay_sinh or "", allow_year_only=True)
    if ns:                           c.ngay_sinh = ns
    nd = parse_date(ngay_chet or "", allow_year_only=True)
    if nd:                           c.ngay_chet = nd
    so = (so_giay_to or "").strip()
    if so:                           c.so_giay_to = so
    nc = parse_date(ngay_cap or "", allow_year_only=True)
    if nc:                           c.ngay_cap  = nc
    dc = (dia_chi or "").strip()
    if dc:                           c.dia_chi   = dc
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        return JSONResponse({"ok": False, "error": "so_giay_to trùng"}, status_code=400)
    db.refresh(c)
    return JSONResponse({"ok": True, "customer": to_customer_json(c)})


@router.post("/create")
def create(
    request: Request,
    ho_ten: Optional[str] = Form(None), gioi_tinh: Optional[str] = Form(None),
    ngay_sinh: Optional[str] = Form(None), ngay_chet: Optional[str] = Form(None),
    so_giay_to: Optional[str] = Form(None), ngay_cap: Optional[str] = Form(None),
    dia_chi: Optional[str] = Form(None),
    db: Session = Depends(get_db)
):
    form = {
        "ho_ten": (ho_ten or "").strip(),
        "gioi_tinh": (gioi_tinh or "").strip(),
        "ngay_sinh": (ngay_sinh or "").strip(),
        "ngay_chet": (ngay_chet or "").strip(),
        "so_giay_to": (so_giay_to or "").strip(),
        "ngay_cap": (ngay_cap or "").strip(),
        "dia_chi": (dia_chi or "").strip(),
    }

    cleaned, field_errors = validate_customer_form(form, db)
    errors = []
    if "so_giay_to" in field_errors:
        errors.append(f"So giay to '{form['so_giay_to']}' da ton tai")

    if field_errors:
        return templates.TemplateResponse("customers/form.html", {
            "request": request, "obj": None,
            "errors": errors, "field_errors": field_errors, "form": form
        })

    c = Customer(**cleaned)
    db.add(c)
    db.commit()
    return RedirectResponse("/customers", status_code=302)


@router.get("/{cid}")
def detail(cid: int, request: Request, db: Session = Depends(get_db)):
    c = db.query(Customer).filter(Customer.id == cid).first()
    if not c:
        raise HTTPException(404)
    return templates.TemplateResponse("customers/detail.html", {"request": request, "obj": c})


@router.get("/{cid}/edit")
def edit_form(cid: int, request: Request, db: Session = Depends(get_db)):
    c = db.query(Customer).filter(Customer.id == cid).first()
    if not c:
        raise HTTPException(404)
    form = {
        "ho_ten": c.ho_ten or "",
        "gioi_tinh": c.gioi_tinh or "",
        "ngay_sinh": format_date_display(c.ngay_sinh),
        "ngay_chet": format_date_display(c.ngay_chet),
        "so_giay_to": c.so_giay_to or "",
        "ngay_cap": format_date_display(c.ngay_cap),
        "dia_chi": c.dia_chi or "",
    }
    return templates.TemplateResponse("customers/form.html", {
        "request": request, "obj": c, "errors": [], "field_errors": {}, "form": form
    })


@router.post("/{cid}/edit")
def edit(
    cid: int, request: Request,
    ho_ten: Optional[str] = Form(None), gioi_tinh: Optional[str] = Form(None),
    ngay_sinh: Optional[str] = Form(None), ngay_chet: Optional[str] = Form(None),
    so_giay_to: Optional[str] = Form(None), ngay_cap: Optional[str] = Form(None),
    dia_chi: Optional[str] = Form(None),
    db: Session = Depends(get_db)
):
    c = db.query(Customer).filter(Customer.id == cid).first()
    if not c:
        raise HTTPException(404)

    form = {
        "ho_ten": (ho_ten or "").strip(),
        "gioi_tinh": (gioi_tinh or "").strip(),
        "ngay_sinh": (ngay_sinh or "").strip(),
        "ngay_chet": (ngay_chet or "").strip(),
        "so_giay_to": (so_giay_to or "").strip(),
        "ngay_cap": (ngay_cap or "").strip(),
        "dia_chi": (dia_chi or "").strip(),
    }

    cleaned, field_errors = validate_customer_form(form, db, current_id=cid)
    errors = []
    if "so_giay_to" in field_errors:
        errors.append(f"So giay to '{form['so_giay_to']}' da ton tai")

    if field_errors:
        return templates.TemplateResponse("customers/form.html", {
            "request": request, "obj": c, "errors": errors, "field_errors": field_errors, "form": form
        })

    c.ho_ten = cleaned["ho_ten"]
    c.gioi_tinh = cleaned["gioi_tinh"]
    c.ngay_sinh = cleaned["ngay_sinh"]
    c.ngay_chet = cleaned["ngay_chet"]
    c.so_giay_to = cleaned["so_giay_to"]
    c.ngay_cap = cleaned["ngay_cap"]
    c.dia_chi = cleaned["dia_chi"]
    db.commit()
    return RedirectResponse(f"/customers/{cid}", status_code=302)


@router.post("/{cid}/delete")
def delete(cid: int, db: Session = Depends(get_db)):
    c = db.query(Customer).filter(Customer.id == cid).first()
    if c:
        db.delete(c)
        db.commit()
    return RedirectResponse("/customers", status_code=302)
