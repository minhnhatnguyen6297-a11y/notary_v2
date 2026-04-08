from __future__ import annotations

import importlib
from importlib import util as importlib_util
import json
import os
import re
import shutil
import time
import warnings
from datetime import date, datetime
from dataclasses import dataclass
from pathlib import Path
from typing import Callable, Optional
import unicodedata

try:
    from dotenv import dotenv_values
except ImportError:  # pragma: no cover
    dotenv_values = None

try:
    from batch_scan import (
        BASE_DIR,
        REGISTRY_DB_PATH,
        connect_registry,
        fetch_registry_records_for_run,
        load_manifest,
        mark_records_uploaded_success,
        now_iso,
        update_registry_record_by_id,
    )
except ImportError:  # pragma: no cover
    from tools.upload_lab.batch_scan import (
        BASE_DIR,
        REGISTRY_DB_PATH,
        connect_registry,
        fetch_registry_records_for_run,
        load_manifest,
        mark_records_uploaded_success,
        now_iso,
        update_registry_record_by_id,
    )

try:
    from uploader_selectors import (
        DEFAULT_GHI_CHU,
        DEFAULT_PHI_CONG_CHUNG,
        DEFAULT_STATUS_TEXT,
        DEFAULT_THU_LAO,
        FORM_FIELD_ORDER,
        FORM_SELECTORS,
        LOGIN_SELECTORS,
        SAVE_BUTTON_SELECTORS,
        VERIFY_FIELDS,
    )
except ImportError:  # pragma: no cover
    from tools.upload_lab.uploader_selectors import (
        DEFAULT_GHI_CHU,
        DEFAULT_PHI_CONG_CHUNG,
        DEFAULT_STATUS_TEXT,
        DEFAULT_THU_LAO,
        FORM_FIELD_ORDER,
        FORM_SELECTORS,
        LOGIN_SELECTORS,
        SAVE_BUTTON_SELECTORS,
        VERIFY_FIELDS,
    )


QUEUE_STATUSES = ("extracted", "upload_failed", "prepared_dry_run", "prepared_partial")
PLAYWRIGHT_MISSING_MESSAGE = (
    "Chua cai Playwright. Scan/Extract van dung duoc. Muon upload, hay chay "
    "'run_ui.bat' de bootstrap tu dong, hoac cai bang "
    "'pip install -r requirements.txt' va 'playwright install chromium'."
)
UPLOAD_LOG_PATH = BASE_DIR / "logs" / "playwright_uploader.log"
DOWNLOADS_DIR = BASE_DIR / "downloads"
UPLOADER_ENV_KEYS = (
    "ND_BASE_URL",
    "ND_LOGIN_URL",
    "ND_CREATE_URL",
    "ND_USERNAME",
    "ND_PASSWORD",
    "ND_STORAGE_STATE_PATH",
    "ND_BROWSER_CHANNEL",
    "ND_MAX_PREPARED_TABS",
    "ND_POST_PREPARE_DELAY_MS",
)
FIELD_VALUE_ALIASES = {
    "nhom_hop_dong": {
        "cam ket - thoa thuan": ["Thoả thuận - Cam kết", "Thỏa thuận - Cam kết", "Cam kết - Thỏa thuận"],
        "thoa thuan - cam ket": ["Thoả thuận - Cam kết", "Thỏa thuận - Cam kết", "Cam kết - Thỏa thuận"],
    }
}


@dataclass
class UploaderSettings:
    base_url: str
    login_url: str
    create_url: str
    username: str
    password: str
    storage_state_path: Path
    browser_channel: str = "chromium"
    max_prepared_tabs: int = 10
    post_prepare_delay_ms: int = 1500


@dataclass
class UploadRecord:
    record_id: int
    run_id: str
    contract_no: str
    status: str
    output_json_path: Path
    source_file: Path
    payload: dict
    upload_form: dict
    missing_fields: list[str]
    raw_row: dict


def _fold_value(text: str) -> str:
    normalized = unicodedata.normalize("NFD", str(text or ""))
    normalized = normalized.replace("\u0111", "d").replace("\u0110", "D")
    folded = "".join(ch for ch in normalized if unicodedata.category(ch) != "Mn").lower()
    return " ".join(folded.split())


def get_field_value_candidates(field_name: str, value: str) -> list[str]:
    raw_value = str(value or "").strip()
    if not raw_value:
        return []

    candidates: list[str] = [raw_value]
    alias_map = FIELD_VALUE_ALIASES.get(field_name, {})
    for alias in alias_map.get(_fold_value(raw_value), []):
        if alias and alias not in candidates:
            candidates.append(alias)
    return candidates


def field_value_matches(field_name: str, expected: str, actual: str) -> bool:
    actual_fold = _fold_value(actual)
    if not actual_fold:
        return False

    for candidate in get_field_value_candidates(field_name, expected):
        candidate_fold = _fold_value(candidate)
        if candidate_fold and candidate_fold in actual_fold:
            return True
    return False


def _default_log(message: str) -> None:
    print(message)


def _default_uploader_env_values(base_dir: Path = BASE_DIR) -> dict[str, str]:
    base_url = "https://congchung.namdinh.gov.vn"
    return {
        "ND_BASE_URL": base_url,
        "ND_LOGIN_URL": base_url,
        "ND_CREATE_URL": f"{base_url}/ho-so-cong-chung/tao-moi-nhanh",
        "ND_USERNAME": "",
        "ND_PASSWORD": "",
        "ND_STORAGE_STATE_PATH": "nd_storage_state.json",
        "ND_BROWSER_CHANNEL": "chromium",
        "ND_MAX_PREPARED_TABS": "10",
        "ND_POST_PREPARE_DELAY_MS": "1500",
    }


def _normalize_env_value(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if len(text) >= 2 and text[0] == text[-1] == '"':
        return text[1:-1].replace('\\"', '"')
    return text


def _resolve_tool_relative_path(base_dir: Path, raw_path: str, *, default_name: str) -> Path:
    raw = str(raw_path or "").strip() or default_name
    normalized = raw.replace("\\", "/")
    if normalized.startswith("tools/upload_lab/"):
        normalized = normalized[len("tools/upload_lab/") :]
    if normalized.startswith("UPLOAD/"):
        normalized = normalized[len("UPLOAD/") :]
    if normalized.startswith("upload/"):
        normalized = normalized[len("upload/") :]
    path = Path(normalized)
    if path.is_absolute():
        return path
    return (base_dir / path).resolve()


def ensure_uploader_env_file(base_dir: Path = BASE_DIR) -> Path:
    base_dir = Path(base_dir)
    env_path = base_dir / ".env"
    env_example_path = base_dir / ".env.example"
    if env_path.exists():
        return env_path

    if env_example_path.exists():
        shutil.copyfile(env_example_path, env_path)
    else:
        save_uploader_env(_default_uploader_env_values(base_dir), base_dir=base_dir)
    return env_path


def read_uploader_env(base_dir: Path = BASE_DIR, *, ensure_exists: bool = False) -> dict[str, str]:
    base_dir = Path(base_dir)
    if ensure_exists:
        ensure_uploader_env_file(base_dir)

    values = _default_uploader_env_values(base_dir)
    env_path = base_dir / ".env"
    if env_path.exists():
        if dotenv_values is not None:
            loaded = dotenv_values(env_path)
            for key, value in loaded.items():
                if key:
                    values[key] = _normalize_env_value(value)
        else:
            for line in env_path.read_text(encoding="utf-8").splitlines():
                stripped = line.strip()
                if not stripped or stripped.startswith("#") or "=" not in stripped:
                    continue
                key, value = stripped.split("=", 1)
                values[key.strip()] = _normalize_env_value(value)

    for key in UPLOADER_ENV_KEYS:
        if key in os.environ:
            values[key] = _normalize_env_value(os.environ[key])
    return values


def _format_env_value(value: object) -> str:
    text = str(value or "")
    if any(ch.isspace() for ch in text) or "#" in text or '"' in text:
        return '"' + text.replace("\\", "\\\\").replace('"', '\\"') + '"'
    return text


def save_uploader_env(values: dict[str, object], *, base_dir: Path = BASE_DIR) -> Path:
    base_dir = Path(base_dir)
    env_path = base_dir / ".env"
    merged = _default_uploader_env_values(base_dir)
    merged.update({key: _normalize_env_value(value) for key, value in values.items() if key in UPLOADER_ENV_KEYS})

    lines = ["# Cau hinh uploader cho tool standalone", ""]
    for key in UPLOADER_ENV_KEYS:
        lines.append(f"{key}={_format_env_value(merged.get(key, ''))}")
    env_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return env_path


def get_uploader_setup_status(base_dir: Path = BASE_DIR) -> dict[str, object]:
    base_dir = Path(base_dir)
    env_path = base_dir / ".env"
    values = read_uploader_env(base_dir)
    storage_state_path = _resolve_tool_relative_path(
        base_dir,
        values.get("ND_STORAGE_STATE_PATH", ""),
        default_name="nd_storage_state.json",
    )
    missing_fields = [
        key
        for key in ("ND_BASE_URL", "ND_LOGIN_URL", "ND_CREATE_URL", "ND_USERNAME", "ND_PASSWORD")
        if not str(values.get(key) or "").strip()
    ]
    storage_state_exists = storage_state_path.exists()
    ready = not missing_fields and storage_state_exists
    if ready:
        message = f"Uploader da san sang. Storage state: {storage_state_path.name}"
    elif missing_fields:
        labels = {
            "ND_BASE_URL": "base url",
            "ND_LOGIN_URL": "login url",
            "ND_CREATE_URL": "create url",
            "ND_USERNAME": "tai khoan",
            "ND_PASSWORD": "mat khau",
        }
        missing_text = ", ".join(labels[key] for key in missing_fields)
        message = f"Can cau hinh uploader lan dau ({missing_text}). Bam 'Cau hinh uploader'."
    else:
        message = "Can dang nhap uploader lan dau de tao nd_storage_state.json. Bam 'Cau hinh uploader'."
    return {
        "env_path": env_path,
        "env_exists": env_path.exists(),
        "values": values,
        "missing_fields": missing_fields,
        "storage_state_path": storage_state_path,
        "storage_state_exists": storage_state_exists,
        "ready": ready,
        "message": message,
    }


def append_log_line(path: Path, message: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    line = f"{now_iso()} {message}\n"
    with path.open("a", encoding="utf-8") as handle:
        handle.write(line)


def probe_playwright_runtime() -> tuple[bool, str]:
    if importlib_util.find_spec("playwright") is None:
        return False, PLAYWRIGHT_MISSING_MESSAGE

    try:
        importlib.import_module("playwright.sync_api")
    except Exception as exc:
        return False, f"Playwright chua san sang: {exc}"

    return True, "Playwright package san sang cho dry-run upload."


def sanitize_contract_no(value: str) -> str:
    text = str(value or "").strip()
    return text.replace("/", "_").replace("\\", "_")


def normalize_web_contract_no(value: str) -> str:
    text = str(value or "").strip().upper()
    if text.endswith("/CCGD"):
        text = text[:-5]
    return text


def normalize_contract_no_for_compare(value: str) -> str:
    text = normalize_web_contract_no(value)
    return re.sub(r"\s+", "", text)


def read_exported_contract_numbers(export_path: Path | str) -> set[str]:
    path = Path(export_path)
    if not path.exists():
        raise FileNotFoundError(f"Khong tim thay file excel doi chieu: {path}")

    try:
        from openpyxl import load_workbook
    except Exception as exc:  # pragma: no cover - dependency wiring
        raise RuntimeError(
            "Thieu openpyxl. Hay chay lai bootstrap hoac cai 'openpyxl' trong moi truong cua tool."
        ) from exc

    contract_nos: set[str] = set()
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message="Workbook contains no default style, apply openpyxl's default",
            category=UserWarning,
        )
        workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        sheet = workbook.active
        for row_index, (cell_value,) in enumerate(
            sheet.iter_rows(min_col=1, max_col=1, values_only=True),
            start=1,
        ):
            raw = str(cell_value or "").strip()
            if not raw:
                continue
            normalized = normalize_contract_no_for_compare(raw)
            header_key = re.sub(r"[^A-Z]", "", _fold_value(raw).upper())
            if row_index == 1 and "SOCONGCHUNG" in header_key:
                continue
            if not normalized:
                continue
            contract_nos.add(normalized)
    finally:
        workbook.close()

    return contract_nos


def split_records_by_existing_contract_nos(
    records: list["UploadRecord"],
    existing_contract_nos: set[str],
) -> tuple[list["UploadRecord"], list["UploadRecord"]]:
    if not existing_contract_nos:
        return list(records), []

    filtered: list[UploadRecord] = []
    duplicates: list[UploadRecord] = []
    for record in records:
        normalized = normalize_contract_no_for_compare(record.contract_no or record.upload_form.get("so_cong_chung", ""))
        if normalized and normalized in existing_contract_nos:
            duplicates.append(record)
        else:
            filtered.append(record)
    return filtered, duplicates


def default_export_from_date() -> str:
    return os.getenv("ND_WEB_EXPORT_FROM_DATE", "01/01/2026").strip() or "01/01/2026"


def default_export_to_date() -> str:
    return date.today().strftime("%d/%m/%Y")


def _parse_export_date_input(value: str) -> str:
    raw = str(value or "").strip()
    if not raw:
        raise ValueError("Ngay xuat so cong chung khong duoc de trong.")

    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(raw, fmt).strftime("%d/%m/%Y")
        except ValueError:
            continue
    raise ValueError(f"Ngay khong hop le: {raw}. Dung DD/MM/YYYY hoac YYYY-MM-DD.")


def load_uploader_settings(base_dir: Path = BASE_DIR) -> UploaderSettings:
    values = read_uploader_env(base_dir)

    base_url = str(values.get("ND_BASE_URL") or "https://congchung.namdinh.gov.vn").rstrip("/")
    login_url = str(values.get("ND_LOGIN_URL") or base_url).strip() or base_url
    create_url = str(values.get("ND_CREATE_URL") or f"{base_url}/ho-so-cong-chung/tao-moi-nhanh").strip()
    storage_state_path = _resolve_tool_relative_path(
        Path(base_dir),
        str(values.get("ND_STORAGE_STATE_PATH") or ""),
        default_name="nd_storage_state.json",
    )
    browser_channel = str(values.get("ND_BROWSER_CHANNEL") or "chromium").strip().lower() or "chromium"
    max_tabs = int(str(values.get("ND_MAX_PREPARED_TABS") or "10"))
    delay_ms = int(str(values.get("ND_POST_PREPARE_DELAY_MS") or "1500"))

    return UploaderSettings(
        base_url=base_url,
        login_url=login_url,
        create_url=create_url,
        username=str(values.get("ND_USERNAME") or "").strip(),
        password=str(values.get("ND_PASSWORD") or "").strip(),
        storage_state_path=storage_state_path,
        browser_channel=browser_channel,
        max_prepared_tabs=max_tabs,
        post_prepare_delay_ms=delay_ms,
    )


def load_upload_queue(
    manifest_path: Path | str,
    *,
    working_dir: Path = BASE_DIR,
    limit: Optional[int] = None,
) -> tuple[dict, list[UploadRecord], int]:
    manifest = load_manifest(manifest_path)
    run_id = str(manifest.get("run_id") or "").strip()
    if not run_id:
        raise ValueError("Manifest khong co run_id")

    conn = connect_registry(working_dir / REGISTRY_DB_PATH.name)
    try:
        rows = fetch_registry_records_for_run(conn, run_id, statuses=QUEUE_STATUSES)
    finally:
        conn.close()

    records: list[UploadRecord] = []
    for row in rows:
        row_dict = dict(row)
        output_json_path = Path(row_dict.get("output_json_path") or "")
        if not output_json_path.exists():
            continue
        payload = json.loads(output_json_path.read_text(encoding="utf-8"))
        upload_form = build_upload_form_data(payload)
        source_file = Path(upload_form["file_hop_dong"])
        missing_fields = identify_missing_fields(upload_form)
        records.append(
            UploadRecord(
                record_id=int(row_dict["id"]),
                run_id=run_id,
                contract_no=str(row_dict.get("contract_no") or upload_form.get("so_cong_chung") or ""),
                status=str(row_dict.get("status") or ""),
                output_json_path=output_json_path,
                source_file=source_file,
                payload=payload,
                upload_form=upload_form,
                missing_fields=missing_fields,
                raw_row=row_dict,
            )
        )

    total_pending = len(records)
    if limit is not None:
        records = records[:limit]
    return manifest, records, total_pending


def finalize_uploaded_records(record_ids: list[int], *, working_dir: Path = BASE_DIR) -> int:
    if not record_ids:
        return 0
    conn = connect_registry(working_dir / REGISTRY_DB_PATH.name)
    try:
        mark_records_uploaded_success(conn, [int(rid) for rid in record_ids])
    finally:
        conn.close()
    return len(record_ids)


def build_upload_form_data(payload: dict) -> dict:
    web_form = dict(payload.get("web_form") or {})
    raw = dict(payload.get("raw") or {})
    return {
        "ten_hop_dong": str(web_form.get("ten_hop_dong") or "").strip(),
        "ngay_cong_chung": str(web_form.get("ngay_cong_chung") or "").strip(),
        "so_cong_chung": normalize_web_contract_no(web_form.get("so_cong_chung") or ""),
        "tinh_trang": DEFAULT_STATUS_TEXT,
        "nhom_hop_dong": str(web_form.get("nhom_hop_dong") or "").strip(),
        "loai_tai_san": str(web_form.get("loai_tai_san") or "").strip(),
        "cong_chung_vien": str(web_form.get("cong_chung_vien") or "").strip(),
        "thu_ky": str(web_form.get("thu_ky") or "").strip(),
        "nguoi_yeu_cau": str(web_form.get("nguoi_yeu_cau") or "").strip(),
        "duong_su": str(web_form.get("duong_su") or "").strip(),
        "tai_san": str(web_form.get("tai_san") or "").strip(),
        "ghi_chu": DEFAULT_GHI_CHU,
        "phi_cong_chung": DEFAULT_PHI_CONG_CHUNG,
        "thu_lao_cong_chung": DEFAULT_THU_LAO,
        "file_hop_dong": str(raw.get("file_goc") or "").strip(),
    }


def identify_missing_fields(upload_form: dict) -> list[str]:
    critical_fields = ("ten_hop_dong", "so_cong_chung", "nhom_hop_dong", "loai_tai_san", "tai_san")
    return [field for field in critical_fields if not str(upload_form.get(field) or "").strip()]


class NamDinhUploaderSession:
    def __init__(
        self,
        settings: UploaderSettings,
        *,
        working_dir: Path = BASE_DIR,
        log_callback: Optional[Callable[[str], None]] = None,
    ):
        self.settings = settings
        self.working_dir = Path(working_dir)
        self._log_callback = log_callback or _default_log
        self.log_path = self.working_dir / "logs" / "playwright_uploader.log"
        self.run_log_path: Path | None = None
        self._playwright = None
        self.browser = None
        self.context = None

    def log(self, message: str) -> None:
        append_log_line(self.log_path, message)
        if self.run_log_path is not None:
            append_log_line(self.run_log_path, message)
        self._log_callback(message)

    @staticmethod
    def _literal_xpath(text: str) -> str:
        if "'" not in text:
            return f"'{text}'"
        if '"' not in text:
            return f'"{text}"'
        parts = text.split("'")
        return "concat(" + ", \"'\", ".join(f"'{part}'" for part in parts) + ")"

    def close(self) -> None:
        if self.context is not None:
            try:
                self.context.close()
            except Exception:
                pass
            self.context = None
        if self.browser is not None:
            try:
                self.browser.close()
            except Exception:
                pass
            self.browser = None
        if self._playwright is not None:
            try:
                self._playwright.stop()
            except Exception:
                pass
            self._playwright = None

    def _import_playwright(self):
        try:
            from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
            from playwright.sync_api import sync_playwright
        except Exception as exc:
            raise RuntimeError(PLAYWRIGHT_MISSING_MESSAGE) from exc
        return sync_playwright, PlaywrightTimeoutError

    def _ensure_context(self) -> None:
        if self.context is not None and self.browser is not None:
            try:
                _ = self.context.pages
                return
            except Exception:
                self.close()

        sync_playwright, _ = self._import_playwright()
        if self._playwright is None:
            self._playwright = sync_playwright().start()

        launch_kwargs = {"headless": False}
        if self.settings.browser_channel and self.settings.browser_channel not in {"", "chromium"}:
            launch_kwargs["channel"] = self.settings.browser_channel

        self.browser = self._playwright.chromium.launch(**launch_kwargs)
        self.settings.storage_state_path.parent.mkdir(parents=True, exist_ok=True)
        context_kwargs = {"accept_downloads": True}
        if self.settings.storage_state_path.exists():
            context_kwargs["storage_state"] = str(self.settings.storage_state_path)
        self.context = self.browser.new_context(**context_kwargs)

    def _save_storage_state(self) -> None:
        if self.context is not None:
            self.context.storage_state(path=str(self.settings.storage_state_path))

    def _locator_from_strategy(self, page, strategy: dict, *, dynamic_text: str = ""):
        stype = strategy.get("type")
        if stype == "css":
            return page.locator(strategy["value"]).first
        if stype == "xpath":
            return page.locator(strategy["value"]).first
        if stype == "text":
            return page.get_by_text(strategy["value"], exact=False).first
        if stype == "role":
            return page.get_by_role(strategy["role"], name=strategy.get("name"), exact=False).first
        if stype == "text_dynamic":
            if dynamic_text:
                return page.get_by_text(dynamic_text, exact=False).first
        return None

    def _label_locator_xpath(self, label_text: str, suffix_xpath: str):
        literal = self._literal_xpath(label_text)
        return (
            f"xpath=((//label[contains(normalize-space(.), {literal})])[1])/{suffix_xpath}"
        )

    def _find_label_locator(self, page, label_text: str):
        literal = self._literal_xpath(label_text)
        candidates = [
            f"xpath=(//label[contains(normalize-space(.), {literal})])[1]",
            f"xpath=(//span[not(*) and contains(normalize-space(.), {literal})])[1]",
            f"xpath=(//p[not(*) and contains(normalize-space(.), {literal})])[1]",
            f"xpath=(//small[not(*) and contains(normalize-space(.), {literal})])[1]",
        ]
        for candidate in candidates:
            locator = page.locator(candidate).first
            if locator.count() > 0:
                return locator
        return None

    def _resolve_control_locator(self, page, field_name: str, *, dynamic_text: str = ""):
        conf = FORM_SELECTORS[field_name]
        for index, strategy in enumerate(conf.get("strategies", []), start=1):
            locator = self._locator_from_strategy(page, strategy, dynamic_text=dynamic_text)
            if locator is not None and locator.count() > 0:
                return locator, f"strategy#{index}:{strategy.get('type')}"

        label = conf.get("label", "")
        kind = conf.get("kind")
        label_locator = self._find_label_locator(page, label)
        if label_locator is None or label_locator.count() == 0:
            return None, f"label_not_found:{label}"

        if kind in {"text", "dropdown"}:
            locator = label_locator.locator(
                "xpath=following::*[self::input and not(@type='hidden') and not(@type='file')][1]"
            ).first
            if locator.count() > 0:
                return locator, "label_following_input"
        if kind == "editor":
            editable = label_locator.locator("xpath=following::*[@contenteditable='true'][1]").first
            if editable.count() > 0:
                return editable, "label_following_editor"
            textarea = label_locator.locator("xpath=following::*[self::textarea][1]").first
            if textarea.count() > 0:
                return textarea, "label_following_textarea"
        if kind == "file":
            locator = label_locator.locator("xpath=following::*[self::input and @type='file'][1]").first
            if locator.count() > 0:
                return locator, "label_following_file"
        return None, f"locator_not_found:{field_name}"

    def _find_control_locator(self, page, field_name: str, *, dynamic_text: str = ""):
        locator, _ = self._resolve_control_locator(page, field_name, dynamic_text=dynamic_text)
        return locator

    def _is_login_page(self, page) -> bool:
        if "dang-nhap" in page.url.lower():
            return True
        username_locator = self._locator_from_strategy(page, LOGIN_SELECTORS["username"][0])
        password_locator = self._locator_from_strategy(page, LOGIN_SELECTORS["password"][0])
        try:
            return bool(
                username_locator
                and password_locator
                and username_locator.count() > 0
                and password_locator.count() > 0
            )
        except Exception:
            return False

    def _perform_login(self, page) -> None:
        if not self.settings.username or not self.settings.password:
            raise RuntimeError("Thieu ND_USERNAME/ND_PASSWORD trong .env de auto login")

        username = None
        password = None
        submit = None
        for strategy in LOGIN_SELECTORS["username"]:
            username = self._locator_from_strategy(page, strategy)
            if username is not None and username.count() > 0:
                break
        for strategy in LOGIN_SELECTORS["password"]:
            password = self._locator_from_strategy(page, strategy)
            if password is not None and password.count() > 0:
                break
        for strategy in LOGIN_SELECTORS["submit"]:
            submit = self._locator_from_strategy(page, strategy)
            if submit is not None and submit.count() > 0:
                break

        if not username or username.count() == 0 or not password or password.count() == 0:
            raise RuntimeError("Khong tim thay form dang nhap de auto login")

        username.fill(self.settings.username)
        password.fill(self.settings.password)
        if submit and submit.count() > 0:
            submit.click()
        else:
            password.press("Enter")
        page.wait_for_timeout(1000)
        page.wait_for_load_state("networkidle")
        self._save_storage_state()

    def ensure_authenticated(self) -> None:
        self._ensure_context()
        page = self.context.new_page()
        try:
            page.goto(self.settings.create_url, wait_until="domcontentloaded")
            page.wait_for_load_state("networkidle")
            if self._is_login_page(page):
                self.log("[UPLOAD] Session het han, dang auto login lai...")
                login_url = self.settings.login_url or self.settings.base_url
                if login_url:
                    page.goto(login_url, wait_until="domcontentloaded")
                    page.wait_for_load_state("networkidle")
                self._perform_login(page)
                page.goto(self.settings.create_url, wait_until="domcontentloaded")
                page.wait_for_load_state("networkidle")
        finally:
            page.close()

    def download_contract_book_export(
        self,
        *,
        from_date: str,
        to_date: str,
        download_dir: Path | None = None,
    ) -> Path:
        from_date_value = _parse_export_date_input(from_date)
        to_date_value = _parse_export_date_input(to_date)
        target_dir = Path(download_dir or (self.working_dir / DOWNLOADS_DIR.name))
        target_dir.mkdir(parents=True, exist_ok=True)

        self.ensure_authenticated()
        page = self.context.new_page()
        try:
            listing_url = f"{self.settings.base_url.rstrip('/')}/ho-so-cong-chung?page=1"
            page.goto(listing_url, wait_until="domcontentloaded")
            page.wait_for_load_state("networkidle")

            export_button = page.get_by_text("Xuất Sổ công chứng", exact=False).first
            if export_button.count() == 0:
                raise RuntimeError("Khong tim thay nut 'Xuất Sổ công chứng' tren trang danh sach.")

            export_button.click()
            dialog = page.get_by_role("dialog").first
            if dialog.count() == 0:
                raise RuntimeError("Khong mo duoc popup chon thoi gian xuat so cong chung.")

            date_inputs = dialog.locator("input[placeholder='dd/mm/yyyy']")
            if date_inputs.count() < 2:
                raise RuntimeError("Khong tim thay du 2 o ngay bat dau/ket thuc trong popup export.")

            date_inputs.nth(0).click()
            date_inputs.nth(0).fill(from_date_value)
            date_inputs.nth(0).press("Tab")
            date_inputs.nth(1).click()
            date_inputs.nth(1).fill(to_date_value)
            date_inputs.nth(1).press("Tab")
            page.wait_for_timeout(300)

            download_button = dialog.get_by_role("button", name="Tải xuống", exact=False).first
            if download_button.count() == 0:
                raise RuntimeError("Khong tim thay nut 'Tải xuống' trong popup export.")

            with page.expect_download(timeout=60000) as download_info:
                download_button.click()
            download = download_info.value
            suggested_name = download.suggested_filename or f"so_cong_chung_{from_date_value.replace('/', '-')}_{to_date_value.replace('/', '-')}.xlsx"
            save_path = target_dir / suggested_name
            if save_path.exists():
                stem = save_path.stem
                suffix = save_path.suffix
                save_path = target_dir / f"{stem}_{now_iso().replace(':', '').replace('-', '')}{suffix}"
            download.save_as(str(save_path))
            self.log(
                f"[UPLOAD][EXPORT] Tai so cong chung tu web OK: {save_path} | from={from_date_value} | to={to_date_value}"
            )
            return save_path
        finally:
            page.close()

    def _fill_text(self, page, field_name: str, value: str) -> bool:
        locator, source = self._resolve_control_locator(page, field_name)
        if locator is None or locator.count() == 0:
            self.log(f"[UPLOAD][FIELD] {field_name}: locator not found ({source})")
            return False
        try:
            locator.click()
            locator.fill(value)
            locator.press("Tab")
            self.log(f"[UPLOAD][FIELD] {field_name}: filled via {source}")
            return True
        except Exception as exc:
            self.log(f"[UPLOAD][FIELD] {field_name}: fill_text failed via {source}: {exc}")
            return False

    def _fill_dropdown(self, page, field_name: str, value: str) -> bool:
        locator, source = self._resolve_control_locator(page, field_name)
        if locator is None or locator.count() == 0:
            self.log(f"[UPLOAD][FIELD] {field_name}: locator not found ({source})")
            return False
        candidates = get_field_value_candidates(field_name, value)
        if not candidates:
            self.log(f"[UPLOAD][FIELD] {field_name}: empty dropdown value")
            return False

        try:
            tag_name = locator.evaluate("(el) => el.tagName.toLowerCase()")
        except Exception:
            tag_name = ""
        current_text = self._read_field_value(page, field_name).strip()
        if tag_name != "input" and field_value_matches(field_name, value, current_text):
            self.log(f"[UPLOAD][FIELD] {field_name}: already set via {source}")
            return True

        last_error = ""
        for candidate in candidates:
            try:
                try:
                    page.keyboard.press("Escape")
                    page.wait_for_timeout(100)
                except Exception:
                    pass

                locator.click()
                locator.fill("")
                locator.type(candidate, delay=25)
            except Exception:
                try:
                    locator.click()
                    locator.fill(candidate)
                except Exception as exc:
                    last_error = str(exc)
                    self.log(
                        f"[UPLOAD][FIELD] {field_name}: fill_dropdown typing failed via {source} with {candidate!r}: {exc}"
                    )
                    continue

            page.wait_for_timeout(300)
            option = page.get_by_role("option", name=candidate, exact=False).first
            if option.count() == 0:
                option = page.get_by_text(candidate, exact=False).first
            if option.count() > 0:
                option.click()
            else:
                locator.press("Enter")
            page.wait_for_timeout(200)

            actual = self._read_field_value(page, field_name).strip()
            if field_value_matches(field_name, candidate, actual):
                self.log(f"[UPLOAD][FIELD] {field_name}: dropdown filled via {source} using {candidate!r}")
                return True

            last_error = f"value_not_stuck expected={candidate!r} actual={actual!r}"
            self.log(
                f"[UPLOAD][FIELD] {field_name}: dropdown candidate {candidate!r} not accepted via {source}; actual={actual!r}"
            )

        self.log(f"[UPLOAD][FIELD] {field_name}: dropdown fill failed via {source}: {last_error}")
        return False

    def _fill_editor(self, page, field_name: str, value: str) -> bool:
        locator, source = self._resolve_control_locator(page, field_name)
        if locator is None or locator.count() == 0:
            self.log(f"[UPLOAD][FIELD] {field_name}: locator not found ({source})")
            return False
        try:
            tag_name = locator.evaluate("(el) => el.tagName.toLowerCase()")
        except Exception:
            tag_name = ""

        if tag_name == "textarea":
            try:
                locator.fill(value)
                self.log(f"[UPLOAD][FIELD] {field_name}: textarea filled via {source}")
                return True
            except Exception as exc:
                self.log(f"[UPLOAD][FIELD] {field_name}: textarea fill failed via {source}: {exc}")
                return False

        try:
            locator.click()
            locator.fill(value)
            locator.press("Tab")
            self.log(f"[UPLOAD][FIELD] {field_name}: editor filled via {source}")
            return True
        except Exception as exc:
            self.log(f"[UPLOAD][FIELD] {field_name}: editor fill failed via {source}: {exc}")
            return False

    def _upload_file(self, page, file_path: Path) -> bool:
        locator, source = self._resolve_control_locator(page, "file_hop_dong", dynamic_text=file_path.name)
        if locator is None or locator.count() == 0:
            self.log(f"[UPLOAD][FIELD] file_hop_dong: locator not found ({source})")
            return False
        try:
            locator.set_input_files(str(file_path))
            page.wait_for_function(
                """(selector) => {
                    const el = document.querySelector(selector);
                    return !!(el && el.files && el.files.length > 0);
                }""",
                arg="#hopdong\\.fileHopdong",
                timeout=5000,
            )
            self.log(f"[UPLOAD][FIELD] file_hop_dong: uploaded via {source} -> {file_path.name}")
            return True
        except Exception as exc:
            success = page.get_by_text(file_path.name, exact=False).count() > 0
            self.log(
                f"[UPLOAD][FIELD] file_hop_dong: upload {'ok' if success else 'failed'} via {source}: {exc}"
            )
            return success

    def _read_field_value(self, page, field_name: str) -> str:
        if field_name == "file_hop_dong":
            locator, _ = self._resolve_control_locator(page, field_name)
            if locator is None or locator.count() == 0:
                return ""
            try:
                return locator.evaluate("(el) => (el.files && el.files.length > 0) ? el.files[0].name : ''")
            except Exception:
                return ""

        locator, _ = self._resolve_control_locator(page, field_name)
        if locator is None or locator.count() == 0:
            return ""
        try:
            tag_name = locator.evaluate("(el) => el.tagName.toLowerCase()")
        except Exception:
            tag_name = ""

        if tag_name in {"input", "textarea", "select"}:
            try:
                return str(locator.input_value()).strip()
            except Exception:
                return ""
        try:
            return str(locator.text_content() or "").strip()
        except Exception:
            return ""

    def _verify_record(self, page, record: UploadRecord) -> tuple[dict, bool]:
        verify_data = {"missing_fields": list(record.missing_fields), "fields": {}}
        partial = bool(record.missing_fields)
        for field_name in VERIFY_FIELDS:
            expected = record.upload_form.get(field_name, "")
            actual = self._read_field_value(page, field_name)
            if field_name == "file_hop_dong":
                success = bool(actual)
                expected_value = Path(expected).name if expected else ""
            else:
                expected_value = str(expected or "").strip()
                if not expected_value:
                    success = False
                elif FORM_SELECTORS.get(field_name, {}).get("kind") == "dropdown":
                    success = field_value_matches(field_name, expected_value, actual)
                else:
                    success = expected_value.lower() in str(actual or "").lower()
            verify_data["fields"][field_name] = {
                "expected": expected_value,
                "actual": actual,
                "success": success,
            }
            self.log(
                f"[UPLOAD][VERIFY] {record.contract_no} {field_name}: "
                f"expected={expected_value!r} actual={actual!r} success={success}"
            )
            if not success:
                partial = True
        return verify_data, partial

    def _prepare_record(self, record: UploadRecord, artifact_dir: Path):
        if not record.source_file.exists():
            self.log(f"[UPLOAD] File goc khong ton tai, tiep tuc khong upload file: {record.source_file}")

        page = self.context.new_page()
        page.goto(self.settings.create_url, wait_until="domcontentloaded")
        page.wait_for_load_state("networkidle")
        if self._is_login_page(page):
            self._perform_login(page)
            page.goto(self.settings.create_url, wait_until="domcontentloaded")
            page.wait_for_load_state("networkidle")

        field_results: dict[str, dict] = {}
        for field_name in FORM_FIELD_ORDER:
            value = str(record.upload_form.get(field_name, "") or "")
            if field_name == "file_hop_dong":
                if value:
                    field_results[field_name] = {"value": Path(value).name, "success": self._upload_file(page, Path(value))}
                continue
            if not value:
                field_results[field_name] = {"value": "", "success": False, "reason": "empty_value"}
                continue
            field_kind = FORM_SELECTORS[field_name]["kind"]
            if field_kind == "text":
                field_results[field_name] = {"value": value, "success": self._fill_text(page, field_name, value)}
            elif field_kind == "dropdown":
                field_results[field_name] = {"value": value, "success": self._fill_dropdown(page, field_name, value)}
            elif field_kind == "editor":
                field_results[field_name] = {"value": value, "success": self._fill_editor(page, field_name, value)}

        verify_data, partial = self._verify_record(page, record)
        screenshot_path = artifact_dir / f"before_save_{sanitize_contract_no(record.contract_no)}.png"
        debug_json_path = artifact_dir / f"debug_{sanitize_contract_no(record.contract_no)}.json"
        page.screenshot(path=str(screenshot_path), full_page=True)
        debug_payload = {
            "contract_no": record.contract_no,
            "source_file": str(record.source_file),
            "page_url": page.url,
            "field_results": field_results,
            "verify": verify_data,
        }
        debug_json_path.write_text(json.dumps(debug_payload, ensure_ascii=False, indent=2), encoding="utf-8")
        if self.settings.post_prepare_delay_ms > 0:
            page.wait_for_timeout(self.settings.post_prepare_delay_ms)

        return {
            "status": "prepared_partial" if partial else "prepared_dry_run",
            "verify_json": json.dumps(verify_data, ensure_ascii=False, indent=2),
            "artifact_dir": str(artifact_dir),
            "screenshot": str(screenshot_path),
            "debug_json": str(debug_json_path),
        }

    def prepare_manifest(
        self,
        manifest_path: Path | str,
        stop_event,
        *,
        exclude_contract_nos: Optional[set[str]] = None,
    ) -> dict:
        manifest, records, total_pending = load_upload_queue(
            manifest_path,
            working_dir=self.working_dir,
        )
        run_id = str(manifest["run_id"])
        filtered_records, duplicate_records = split_records_by_existing_contract_nos(
            records,
            set(exclude_contract_nos or set()),
        )
        records = filtered_records[: self.settings.max_prepared_tabs]
        filtered_pending = len(filtered_records)

        if not records:
            duplicate_message = ""
            if duplicate_records:
                duplicate_message = f" Da loai {len(duplicate_records)} ho so trung so cong chung tren web."
            return {
                "run_id": run_id,
                "prepared_count": 0,
                "total_pending": total_pending,
                "remaining": 0,
                "artifact_dir": "",
                "excluded_duplicates": len(duplicate_records),
                "message": f"Khong con ho so nao can chuan bi.{duplicate_message}",
            }

        artifact_dir = self.working_dir / "upload_runs" / f"{now_iso().replace(':', '').replace('-', '')}_{run_id}"
        artifact_dir.mkdir(parents=True, exist_ok=True)
        self.run_log_path = artifact_dir / "dry_run_trace.log"
        self.log(f"[UPLOAD] Artifact dir: {artifact_dir}")
        self.ensure_authenticated()

        prepared_count = 0
        errors = []
        conn = connect_registry(self.working_dir / REGISTRY_DB_PATH.name)
        try:
            for record in records:
                if stop_event.is_set():
                    break
                self.log(f"[UPLOAD] Dang chuan bi {record.contract_no} (record #{record.record_id})")
                try:
                    result = self._prepare_record(record, artifact_dir)
                    update_registry_record_by_id(
                        conn,
                        record.record_id,
                        status=result["status"],
                        reason="Dry-run prepared",
                        last_error="",
                        prepared_at=now_iso(),
                        artifact_dir=result["artifact_dir"],
                        verify_json=result["verify_json"],
                    )
                    prepared_count += 1
                    self.log(
                        f"[UPLOAD] {record.contract_no}: {result['status']} -> {result['screenshot']} | debug={result['debug_json']}"
                    )
                except Exception as exc:
                    errors.append({"record_id": record.record_id, "contract_no": record.contract_no, "error": str(exc)})
                    update_registry_record_by_id(
                        conn,
                        record.record_id,
                        status="upload_failed",
                        last_error=str(exc),
                        reason="Uploader dry-run failed",
                        artifact_dir=str(artifact_dir),
                    )
                    self.log(f"[UPLOAD] Loi {record.contract_no}: {exc}")
            summary = {
                "run_id": run_id,
                "prepared_count": prepared_count,
                "total_pending": filtered_pending,
                "remaining": max(filtered_pending - prepared_count, 0),
                "artifact_dir": str(artifact_dir),
                "excluded_duplicates": len(duplicate_records),
                "errors": errors,
                "message": (
                    f"Da chuan bi {prepared_count}/{filtered_pending} ho so."
                    + (f" Da loai {len(duplicate_records)} ho so trung tren web." if duplicate_records else "")
                    + " Hay ra soat, luu, finalize roi chay tiep chunk sau."
                ),
            }
            (artifact_dir / "upload_manifest.json").write_text(
                json.dumps(summary, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
            return summary
        finally:
            self.run_log_path = None
            conn.close()


def download_contract_book_export(
    *,
    from_date: str,
    to_date: str,
    working_dir: Path = BASE_DIR,
    settings: Optional[UploaderSettings] = None,
    log_callback: Optional[Callable[[str], None]] = None,
) -> Path:
    session = NamDinhUploaderSession(
        settings or load_uploader_settings(working_dir),
        working_dir=working_dir,
        log_callback=log_callback,
    )
    try:
        return session.download_contract_book_export(from_date=from_date, to_date=to_date)
    finally:
        session.close()
